import pandas as pd
import numpy as np
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side, Color
from openpyxl.utils import get_column_letter
from openpyxl.chart import LineChart, BarChart, PieChart, Reference
from openpyxl.chart.label import DataLabelList
import warnings
warnings.filterwarnings('ignore')

class AdvancedDairyAnalyzer:
    def __init__(self, file_path):
        """Initialize the analyzer with data preprocessing"""
        try:
            self.df = pd.read_csv(file_path)    #بجای file_path مسیر فایل دیتاست را قرار دهید
            self._preprocess_data()
            print("داده‌ها با موفقیت خوانده شد.")
            print(f"تعداد رکوردها: {len(self.df)}")
        except Exception as e:
            print(f"خطا در خواندن فایل: {str(e)}")
            raise
        
    def _preprocess_data(self):
        """پردازش و تمیزسازی داده‌ها"""
        # تبدیل ستون‌های تاریخ
        date_columns = ['Date', 'Production Date', 'Expiration Date']
        for col in date_columns:
            self.df[col] = pd.to_datetime(self.df[col])
            
        # محاسبه فیلدهای جدید
        self.df['Profit Margin'] = ((self.df['Price per Unit (sold)'] - self.df['Price per Unit']) / 
                                  self.df['Price per Unit']) * 100
        self.df['Stock Duration'] = (self.df['Expiration Date'] - self.df['Production Date']).dt.days
        self.df['Days to Expiry'] = (self.df['Expiration Date'] - self.df['Date']).dt.days
        
        # محاسبه شاخص‌های فصلی
        self.df['Month'] = self.df['Date'].dt.month
        self.df['Season'] = pd.cut(self.df['Date'].dt.month, bins=[0,3,6,9,12], 
                                 labels=['بهار', 'تابستان', 'پاییز', 'زمستان'])
        
        # محاسبه شاخص‌های کارایی
        self.df['Revenue per Cow'] = self.df['Approx. Total Revenue(INR)'] / self.df['Number of Cows']
        self.df['Revenue per Acre'] = self.df['Approx. Total Revenue(INR)'] / self.df['Total Land Area (acres)']

    def sales_trend_analysis(self):
        """تحلیل روند فروش"""
        monthly_sales = self.df.groupby([self.df['Date'].dt.year, 
                                       self.df['Date'].dt.month]).agg({
            'Quantity Sold (liters/kg)': 'sum',
            'Approx. Total Revenue(INR)': 'sum',
            'Profit Margin': 'mean'
        }).round(2)
        
        monthly_sales.index = [f"{year}-{month:02d}" for year, month in monthly_sales.index]
        monthly_sales.columns = ['حجم فروش', 'درآمد', 'حاشیه سود']
        return monthly_sales

    def product_performance(self):
        """تحلیل عملکرد محصولات"""
        product_stats = self.df.groupby('Product Name').agg({
            'Quantity Sold (liters/kg)': ['sum', 'mean', 'std'],
            'Price per Unit (sold)': ['mean', 'min', 'max'],
            'Profit Margin': ['mean', 'min', 'max'],
            'Approx. Total Revenue(INR)': 'sum'
        })
        
        product_stats.columns = [
            'مجموع فروش',
            'میانگین فروش',
            'انحراف معیار فروش',
            'میانگین قیمت',
            'حداقل قیمت',
            'حداکثر قیمت',
            'میانگین حاشیه سود',
            'حداقل حاشیه سود',
            'حداکثر حاشیه سود',
            'مجموع درآمد'
        ]
        
        # محاسبه رتبه و سهم بازار
        product_stats['رتبه'] = product_stats['مجموع درآمد'].rank(ascending=False)
        total_revenue = product_stats['مجموع درآمد'].sum()
        product_stats['سهم بازار (%)'] = (product_stats['مجموع درآمد'] / total_revenue * 100).round(2)
        
        return product_stats.sort_values('مجموع درآمد', ascending=False)

    def farm_efficiency_analysis(self):
        """تحلیل کارایی مزارع"""
        farm_stats = self.df.groupby(['Location', 'Farm Size']).agg({
            'Number of Cows': 'mean',
            'Total Land Area (acres)': 'mean',
            'Revenue per Cow': 'mean',
            'Revenue per Acre': 'mean',
            'Quantity Sold (liters/kg)': 'sum',
            'Approx. Total Revenue(INR)': 'sum'
        }).round(2)
        
        farm_stats.columns = [
            'متوسط تعداد گاو',
            'متوسط مساحت (هکتار)',
            'درآمد سرانه هر گاو',
            'درآمد سرانه هر هکتار',
            'مجموع تولید',
            'مجموع درآمد'
        ]
        
        return farm_stats

    def inventory_analysis(self):
        """تحلیل موجودی"""
        inventory_stats = self.df.groupby('Product Name').agg({
            'Quantity in Stock (liters/kg)': ['sum', 'mean'],
            'Minimum Stock Threshold (liters/kg)': 'mean',
            'Days to Expiry': ['mean', 'min'],
            'Quantity Sold (liters/kg)': lambda x: x.mean() * 30
        })
        
        inventory_stats.columns = [
            'موجودی کل',
            'میانگین موجودی',
            'حداقل موجودی مجاز',
            'میانگین روزهای تا انقضا',
            'حداقل روزهای تا انقضا',
            'میانگین فروش ماهانه'
        ]
        
        # محاسبه شاخص‌های ریسک
        inventory_stats['نسبت موجودی به فروش'] = (
            inventory_stats['موجودی کل'] / inventory_stats['میانگین فروش ماهانه']
        ).round(2)
        
        inventory_stats['شاخص ریسک'] = (
            inventory_stats['نسبت موجودی به فروش'] * 
            (30 / inventory_stats['میانگین روزهای تا انقضا'])
        ).round(2)
        
        return inventory_stats

    def customer_analysis(self):
        """تحلیل مشتریان"""
        customer_stats = self.df.groupby('Customer Location').agg({
            'Quantity Sold (liters/kg)': ['sum', 'mean'],
            'Approx. Total Revenue(INR)': ['sum', 'mean'],
            'Profit Margin': 'mean'
        })
        
        customer_stats.columns = [
            'مجموع فروش',
            'میانگین فروش',
            'مجموع درآمد',
            'میانگین درآمد',
            'میانگین حاشیه سود'
        ]
        
        # محاسبه سهم بازار
        total_revenue = customer_stats['مجموع درآمد'].sum()
        customer_stats['سهم بازار (%)'] = (
            customer_stats['مجموع درآمد'] / total_revenue * 100
        ).round(2)
        
        return customer_stats.sort_values('مجموع درآمد', ascending=False)

    def generate_excel_report(self, output_path='dairy_analysis_report.xlsx'):
        """تولید گزارش اکسل با فرمت‌بندی پیشرفته"""
        try:
            with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
                workbook = writer.book
                
                # تنظیمات استایل
                header_style = {
                    'font': Font(name='B Nazanin', bold=True, size=12, color='FFFFFF'),
                    'fill': PatternFill(start_color='366092', end_color='366092', fill_type='solid'),
                    'alignment': Alignment(horizontal='center', vertical='center', wrap_text=True),
                    'border': Border(
                        left=Side(style='thin'),
                        right=Side(style='thin'),
                        top=Side(style='thin'),
                        bottom=Side(style='thin')
                    )
                }
                
                # ۱. خلاصه مدیریتی
                self._create_executive_summary(writer)
                
                # ۲. روند فروش
                sales_trend = self.sales_trend_analysis()
                sales_trend.to_excel(writer, sheet_name='روند فروش')
                self._format_sheet(writer.sheets['روند فروش'], header_style)
                self._add_sales_chart(writer.sheets['روند فروش'], sales_trend)
                
                # ۳. تحلیل محصولات
                products = self.product_performance()
                products.to_excel(writer, sheet_name='تحلیل محصولات')
                self._format_sheet(writer.sheets['تحلیل محصولات'], header_style)
                
                # ۴. تحلیل کارایی مزارع
                farms = self.farm_efficiency_analysis()
                farms.to_excel(writer, sheet_name='کارایی مزارع')
                self._format_sheet(writer.sheets['کارایی مزارع'], header_style)
                
                # ۵. تحلیل موجودی
                inventory = self.inventory_analysis()
                inventory.to_excel(writer, sheet_name='تحلیل موجودی')
                self._format_sheet(writer.sheets['تحلیل موجودی'], header_style)
                
                # ۶. تحلیل مشتریان
                customers = self.customer_analysis()
                customers.to_excel(writer, sheet_name='تحلیل مشتریان')
                self._format_sheet(writer.sheets['تحلیل مشتریان'], header_style)
                
                # تنظیم عرض ستون‌ها برای تمام شیت‌ها
                for sheet in workbook.sheetnames:
                    self._adjust_column_width(writer.sheets[sheet])
                
                print(f"گزارش با موفقیت در مسیر {output_path} ذخیره شد.")
                
        except Exception as e:
            print(f"خطا در ایجاد گزارش اکسل: {str(e)}")
            raise

    def _create_executive_summary(self, writer):
        """ایجاد خلاصه مدیریتی"""
        # محاسبه شاخص‌های کلیدی
        total_revenue = self.df['Approx. Total Revenue(INR)'].sum()
        total_quantity = self.df['Quantity Sold (liters/kg)'].sum()
        avg_margin = self.df['Profit Margin'].mean()
        total_farms = len(self.df['Location'].unique())
        total_products = len(self.df['Product Name'].unique())
        
        summary_data = pd.DataFrame({
            'شاخص': [
                'مجموع درآمد (INR)',
                'مجموع فروش (لیتر/کیلوگرم)',
                'میانگین حاشیه سود (%)',
                'تعداد مزارع فعال',
                'تعداد محصولات',
                'میانگین روزهای ماندگاری محصول'
            ],
            'مقدار': [
                f"{total_revenue:,.2f}",
                f"{total_quantity:,.2f}",
                f"{avg_margin:.2f}",
                str(total_farms),
                str(total_products),
                f"{self.df['Days to Expiry'].mean():.1f}"
            ]
        })
        
        summary_data.to_excel(writer, sheet_name='خلاصه مدیریتی', index=False)
        sheet = writer.sheets['خلاصه مدیریتی']
        
        # فرمت‌بندی
        for row in sheet.iter_rows(min_row=1, max_row=sheet.max_row):
            for cell in row:
                cell.alignment = Alignment(horizontal='center')
                if cell.row == 1:
                    cell.font = Font(bold=True, size=12, color='FFFFFF')
                    cell.fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
                else:
                    cell.font = Font(size=11)

    def _format_sheet(self, sheet, header_style):
        """فرمت‌بندی شیت‌های اکسل"""
        # فرمت‌بندی هدرها
        for cell in sheet[1]:
            if cell.value:
                for key, value in header_style.items():
                    setattr(cell, key, value)
        
        # فرمت‌بندی سلول‌های داده
        for row in sheet.iter_rows(min_row=2):
            for cell in row:
                if cell.value is not None:
                    cell.alignment = Alignment(horizontal='center')
                    cell.font = Font(size=11)
                    try:
                        if isinstance(cell.value, (int, float)):
                            cell.number_format = '#,##0.00'
                    except:
                        pass

    def _adjust_column_width(self, sheet):
        """تنظیم عرض ستون‌ها"""
        for column in sheet.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            
            adjusted_width = max_length + 4
            sheet.column_dimensions[column_letter].width = adjusted_width

    def _add_sales_chart(self, sheet, data):
        """اضافه کردن نمودار روند فروش"""
        chart = LineChart()
        chart.title = "روند فروش ماهانه"
        chart.style = 2
        chart.height = 15
        chart.width = 30
        
        # اضافه کردن داده‌ها به نمودار
        data_ref = Reference(sheet, min_col=2, min_row=1, max_row=len(data)+1, max_col=2)
        cats_ref = Reference(sheet, min_col=1, min_row=2, max_row=len(data)+1)
        
        chart.add_data(data_ref, titles_from_data=True)
        chart.set_categories(cats_ref)
        
        # تنظیمات نمودار
        chart.x_axis.title = "ماه"
        chart.y_axis.title = "مقدار فروش"
        chart.legend = None
        
        # اضافه کردن برچسب‌های داده
        chart.dataLabels = DataLabelList()
        chart.dataLabels.showVal = True
        
        # قرار دادن نمودار در شیت
        sheet.add_chart(chart, "H2")

def main():
    """تابع اصلی برای اجرای تحلیل"""
    try:
        # ایجاد نمونه از کلاس تحلیلگر
        analyzer = AdvancedDairyAnalyzer('dairy_dataset.csv')
        
        # تولید گزارش اکسل
        analyzer.generate_excel_report()
        
        print("تحلیل با موفقیت انجام شد!")
        
    except Exception as e:
        print(f"خطا در اجرای برنامه: {str(e)}")
        raise

if __name__ == "__main__":
    main()